"""
dashboard_router.py — Staff KPI Dashboard

Mounted into the main FastAPI app in sync_service.py via:
    app.include_router(create_dashboard_router(service))

Routes:
    GET  /dashboard              → Interactive HTML dashboard page
    GET  /api/dashboard/stats    → JSON KPI data (consumed by the page via fetch())
"""

from __future__ import annotations

import io
import json
import os
import secrets
import time
from datetime import date, datetime, timedelta
from typing import Any, Dict, List

from fastapi import APIRouter, Form, Query, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

# ── Staff sheet cache (avoid re-fetching on every dashboard refresh) ─────────
_staff_cache: Dict = {"data": None, "ts": 0.0}
_STAFF_CACHE_TTL = 300  # seconds

# ── Stats result cache: keyed by (date_from, date_to) ─────────────────────
# Avoids hammering AMO on every auto-refresh or repeated date query.
# TTL: 60 s when today is in the range (data is live), 300 s for past-only ranges.
_stats_cache: Dict = {}  # (date_from, date_to) → {"ts": float, "data": dict}
_leads_cache: Dict = {}  # (date_from, date_to) → {"ts": float, "leads": list}

# ── Session store ─────────────────────────────────────────────────
# token → {"username": str, "created_at": float}
_sessions: Dict[str, Dict] = {}
_SESSION_TTL = int(os.getenv("DASHBOARD_SESSION_TTL", str(8 * 3600)).strip("'\""))  # default 8 h


def _load_admins() -> List[Dict[str, str]]:
    """Load admin credentials from env.

    Priority:
    1. DASHBOARD_ADMINS_JSON  = '[{"username":"alice","password":"s3cr3t"}, ...]'
    2. DASHBOARD_ADMIN_USERNAME + DASHBOARD_ADMIN_PASSWORD  (single admin fallback)
    """
    raw = os.getenv("DASHBOARD_ADMINS_JSON", "").strip()
    if raw:
        try:
            admins = json.loads(raw)
            if isinstance(admins, list) and admins:
                return admins
        except Exception:
            pass
    u = os.getenv("DASHBOARD_ADMIN_USERNAME", "admin").strip()
    p = os.getenv("DASHBOARD_ADMIN_PASSWORD", "").strip()
    if p:
        return [{"username": u, "password": p}]
    return []  # no credentials configured — login will always fail


def _check_session(token: str) -> str | None:
    """Return username if token is valid and not expired, else None."""
    if not token:
        return None
    entry = _sessions.get(token)
    if not entry:
        return None
    if time.time() - entry["created_at"] > _SESSION_TTL:
        _sessions.pop(token, None)
        return None
    return entry["username"]

# ── Status display names that count as a confirmed order for KPI ─────────────
# Includes every stage at or beyond "Заказ" — the order has been placed.
ZAKAS_DISPLAY_NAMES: set[str] = {"Заказ", "В процессе", "У курера", "Успешно"}

# Statuses that count as rejection
OTKAZ_DISPLAY_NAMES: set[str] = {"Отказ", "Закрыто и не реализовано"}

# Statuses that count as consideration (lead is thinking/hesitating)
DUMKA_DISPLAY_NAMES: set[str] = {"Раздумье"}


def _norm(name: str) -> str:
    """Lowercase + collapse whitespace for fuzzy name matching."""
    return " ".join(name.lower().split())


# ─────────────────────────────────────────────────────────────────────────────
# Factory – call once from sync_service.py, passing the live SyncService.
# ─────────────────────────────────────────────────────────────────────────────

def create_dashboard_router(service) -> APIRouter:
    router = APIRouter()

    # ── Auth helpers ──────────────────────────────────────────────────────────
    def _user(request: Request) -> str | None:
        return _check_session(request.cookies.get("dash_token", ""))

    def _login_response(error: str = "") -> HTMLResponse:
        err_block = (
            f'<div class="err" style="margin-bottom:16px">{error}</div>'
            if error else ""
        )
        return HTMLResponse(_LOGIN_HTML.replace("{error_block}", err_block))

    # ── Login / Logout ────────────────────────────────────────────────────────
    @router.get("/login", response_class=HTMLResponse, tags=["auth"])
    def login_page(request: Request):
        if _user(request):
            return RedirectResponse("/dashboard", status_code=302)
        return _login_response()

    @router.post("/login", tags=["auth"])
    def login_submit(
        request: Request,
        username: str = Form(...),
        password: str = Form(...),
    ):
        admins = _load_admins()
        if not admins:
            return _login_response("Нет настроенных аккаунтов. Установите DASHBOARD_ADMIN_PASSWORD в .env")
        match = any(
            a.get("username", "").strip() == username.strip()
            and a.get("password", "") == password
            for a in admins
        )
        if not match:
            return _login_response("Неверный логин или пароль.")
        token = secrets.token_hex(32)
        _sessions[token] = {"username": username.strip(), "created_at": time.time()}
        resp = RedirectResponse("/dashboard", status_code=302)
        resp.set_cookie(
            "dash_token", token,
            max_age=_SESSION_TTL, httponly=True, samesite="lax",
        )
        return resp

    @router.get("/logout", tags=["auth"])
    def logout(request: Request):
        token = request.cookies.get("dash_token", "")
        _sessions.pop(token, None)
        resp = RedirectResponse("/login", status_code=302)
        resp.delete_cookie("dash_token")
        return resp

    # ── HTML page ─────────────────────────────────────────────────────────────
    @router.get("/dashboard", response_class=HTMLResponse, tags=["dashboard"])
    def dashboard_page(request: Request) -> HTMLResponse:
        if not _user(request):
            return RedirectResponse("/login", status_code=302)
        return HTMLResponse(_DASHBOARD_HTML)

    # ── JSON stats API ────────────────────────────────────────────────────────
    _snap_dates_cache: Dict = {"ts": 0.0, "data": []}

    @router.get("/api/dashboard/stats", tags=["dashboard"])
    def dashboard_stats(
        date_from:  str = Query(default="", description="YYYY-MM-DD, defaults to today"),
        date_to:    str = Query(default="", description="YYYY-MM-DD, defaults to today"),
        group:      str = Query(default="", description="Group filter"),
        staff_code: str = Query(default="", description="Filter by staff code"),
        force:      int = Query(default=0,  description="Set to 1 to bypass cache"),
    ) -> Dict[str, Any]:
        import traceback
        _empty = {
            "groups": {}, "date_from": date_from, "date_to": date_to,
            "is_live": False, "total_consul": 0, "total_zakas": 0,
            "total_summa": 0, "total_uspeshka": 0, "total_uspeshka_summa": 0,
            "avg_zakaz_conv": 0.0, "avg_uspeshka_conv": 0.0,
        }
        try:
            today = date.today().strftime("%Y-%m-%d")
            if not date_from: date_from = today
            if not date_to:   date_to   = today
            yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")

            # ── Cache ──────────────────────────────────────────────────────────
            cache_key = (date_from, date_to, group, staff_code)
            has_today = (date_from <= today <= date_to)
            cache_ttl = 60 if has_today else 300
            cached    = _stats_cache.get(cache_key)
            if cached and not force and (time.monotonic() - cached["ts"]) < cache_ttl:
                return cached["data"]

            # ── 1. Load Staff sheet ───────────────────────────────────────────
            staff_by_code: Dict[str, Dict] = {}
            now_ts = time.monotonic()
            if _staff_cache["data"] is not None and (now_ts - _staff_cache["ts"]) < _STAFF_CACHE_TTL:
                staff_by_code = _staff_cache["data"]
            else:
                try:
                    staff_by_code = service.get_staff_list()
                    _staff_cache["data"] = staff_by_code
                    _staff_cache["ts"]   = now_ts
                except Exception as exc:
                    print(f"[DASHBOARD] Could not load Staff sheet: {exc}")
                    if _staff_cache["data"] is not None:
                        staff_by_code = _staff_cache["data"]

            # ── 2. Smart data source ───────────────────────────────────────────
            # Past days with snapshots → manager_snapshots (0 AMO API calls)
            # Today without a snapshot → live kpi_events
            merged: Dict[str, Dict] = {}  # staff_code → aggregated row

            def _add_kpi_row(code: str, row: Dict) -> None:
                if code not in merged:
                    merged[code] = {
                        "consul": 0, "zakas": 0, "summa": 0.0,
                        "uspeshka": 0, "uspeshka_summa": 0.0,
                        "otkaz": 0, "dumka": 0,
                    }
                m = merged[code]
                m["consul"]         += int(row.get("consul", 0))
                m["zakas"]          += int(row.get("zakas", 0))
                m["summa"]          += float(row.get("summa", 0))
                m["uspeshka"]       += int(row.get("uspeshka", 0))
                m["uspeshka_summa"] += float(row.get("uspeshka_summa", 0))
                m["otkaz"]          += int(row.get("otkaz", 0))
                m["dumka"]          += int(row.get("dumka", 0))

            # Past days with snapshots
            if date_from <= yesterday:
                snap_to   = min(date_to, yesterday)
                snap_rows = service.kpi_store.get_manager_stats_snapshot(date_from, snap_to)
                for row in snap_rows:
                    _add_kpi_row(str(row["staff_code"]), row)

            # Today (live kpi_events)
            is_live = False
            if date_to >= today:
                is_live   = not service.kpi_store.has_snapshot(today)
                live_rows = service.kpi_store.get_staff_stats(today, today)
                for row in live_rows:
                    _add_kpi_row(str(row["staff_code"]), row)

            # ── 3. Resolve staff, apply filters, compute KPIs ─────────────────
            rows_out: List[Dict] = []
            skipped_unknown = 0
            for code, agg in merged.items():
                if staff_code and code != staff_code.strip():
                    continue
                staff_info = (
                    staff_by_code.get(code)
                    or staff_by_code.get(code.lstrip("0"))
                    or staff_by_code.get(code.zfill(4))
                )
                if not staff_info:
                    skipped_unknown += 1
                    continue
                g = staff_info["group"]
                if group and g.upper() != group.upper():
                    continue
                consul         = agg["consul"]
                zakas          = agg["zakas"]
                summa          = agg["summa"]
                uspeshka       = agg["uspeshka"]
                uspeshka_summa = agg["uspeshka_summa"]
                otkaz          = agg["otkaz"]
                dumka          = agg["dumka"]
                zakaz_conv    = round(zakas / consul * 100, 1) if consul else 0.0
                uspeshka_conv = round(uspeshka_summa / summa * 100, 1) if summa else 0.0
                rows_out.append({
                    "code":           code,
                    "name":           staff_info["full_name"],
                    "group":          g,
                    "consul":         consul,
                    "zakas":          zakas,
                    "summa":          int(summa),
                    "uspeshka":       uspeshka,
                    "uspeshka_summa": int(uspeshka_summa),
                    "otkaz":          otkaz,
                    "dumka":          dumka,
                    "zakaz_conv":     zakaz_conv,
                    "uspeshka_conv":  uspeshka_conv,
                })

            # ── 4. Sort + number ──────────────────────────────────────────────
            rows_out.sort(key=lambda x: (-x["summa"], x["name"]))

            # ── 5. Group by Группа ────────────────────────────────────────────
            groups: Dict[str, List] = {}
            for r in rows_out:
                g = r["group"] or "—"
                groups.setdefault(g, []).append(r)
            for g_rows in groups.values():
                for i, row in enumerate(g_rows, 1):
                    row["num"] = i

            # ── 6. Totals ─────────────────────────────────────────────────────
            all_consul         = sum(r["consul"]         for r in rows_out)
            all_zakas          = sum(r["zakas"]          for r in rows_out)
            all_summa          = sum(r["summa"]          for r in rows_out)
            all_uspeshka       = sum(r["uspeshka"]       for r in rows_out)
            all_uspeshka_summa = sum(r["uspeshka_summa"] for r in rows_out)
            avg_zakaz_conv    = round(all_zakas / all_consul * 100, 1) if all_consul else 0.0
            avg_uspeshka_conv = round(all_uspeshka_summa / all_summa * 100, 1) if all_summa else 0.0

            result = {
                "date_from":            date_from,
                "date_to":              date_to,
                "is_live":              is_live,
                "total_consul":         all_consul,
                "total_zakas":          all_zakas,
                "total_summa":          all_summa,
                "total_uspeshka":       all_uspeshka,
                "total_uspeshka_summa": all_uspeshka_summa,
                "avg_zakaz_conv":       avg_zakaz_conv,
                "avg_uspeshka_conv":    avg_uspeshka_conv,
                "skipped_unknown":      skipped_unknown,
                "groups":               groups,
            }
            _stats_cache[cache_key] = {"ts": time.monotonic(), "data": result}
            return result

        except Exception as exc:
            print(f"[DASHBOARD] Error in stats endpoint: {traceback.format_exc()}")
            return {**_empty, "error": str(exc), "date_from": date_from, "date_to": date_to}

    @router.get("/api/dashboard/snapshot-dates", tags=["dashboard"])
    def snapshot_dates_endpoint(request: Request) -> Dict[str, Any]:
        """Return available snapshot dates, cached 60 s."""
        if not _user(request):
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Unauthorized")
        now = time.monotonic()
        if now - _snap_dates_cache["ts"] < 60:
            return {"dates": _snap_dates_cache["data"]}
        dates = service.kpi_store.get_available_snapshot_dates()
        _snap_dates_cache["ts"]   = now
        _snap_dates_cache["data"] = dates
        return {"dates": dates}

    @router.post("/api/manager/snapshot", tags=["dashboard"])
    async def trigger_snapshot(request: Request) -> Dict[str, Any]:
        """Manually trigger a nightly snapshot for the given date. Requires auth."""
        if not _user(request):
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Unauthorized")
        try:
            payload = await request.json()
        except Exception:
            payload = {}
        snap_date = (payload.get("date") or "").strip()
        if not snap_date:
            return {"error": "date field required (YYYY-MM-DD)"}
        import threading as _th
        _th.Thread(
            target=service.run_nightly_snapshot,
            args=(snap_date,),
            daemon=True,
            name="manual-snapshot",
        ).start()
        return {"status": "started", "date": snap_date}

    # ── Monthly report endpoint ───────────────────────────────────────────────
    @router.get("/api/dashboard/monthly-report", tags=["dashboard"])
    def monthly_report(
        request: Request,
        month:   str = Query(default="", description="YYYY-MM, defaults to current month"),
        group:   str = Query(default="", description="Group filter: A / B / C / D"),
    ) -> Dict[str, Any]:
        """Return per-staff KPI aggregation for an entire month.

        Because data is stored event-by-event in SQLite, this endpoint is
        fast and does not require any AMO API calls.

        Returns the same structure as /api/dashboard/stats plus a daily breakdown
        suitable for building monthly salary reports.
        """
        import traceback
        try:
            if not month:
                from datetime import date as _date
                month = _date.today().strftime("%Y-%m")

            # Resolve full month date range
            from datetime import date as _date
            y, m = int(month[:4]), int(month[5:7])
            date_from = f"{month}-01"
            if m == 12:
                last_day = _date(y + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = _date(y, m + 1, 1) - timedelta(days=1)
            date_to = last_day.strftime("%Y-%m-%d")

            # ── Load staff for name/group resolution ─────────────────────────
            staff_by_code: Dict[str, Dict] = {}
            now_ts = time.monotonic()
            if _staff_cache["data"] is not None and (now_ts - _staff_cache["ts"]) < _STAFF_CACHE_TTL:
                staff_by_code = _staff_cache["data"]
            else:
                try:
                    staff_by_code = service.get_staff_list()
                    _staff_cache["data"] = staff_by_code
                    _staff_cache["ts"]   = now_ts
                except Exception as exc:
                    print(f"[DASHBOARD] Could not load Staff sheet: {exc}")
                    if _staff_cache["data"] is not None:
                        staff_by_code = _staff_cache["data"]

            # ── Monthly totals from KPI store ────────────────────────────────
            kpi_rows = service.kpi_store.get_staff_stats(date_from, date_to)

            # ── Daily breakdown ───────────────────────────────────────────────
            daily_rows = service.kpi_store.get_daily_breakdown(date_from, date_to)
            # daily_map: staff_code → {date → {consul, zakas, dumka, summa, uspeshka, uspeshka_summa}}
            daily_map: Dict[str, Dict] = {}
            for dr in daily_rows:
                code = str(dr["staff_code"])
                d    = dr["event_date"]
                daily_map.setdefault(code, {})[d] = {
                    "consul":         int(dr["consul"]),
                    "zakas":          int(dr["zakas"]),
                    "dumka":          int(dr["dumka"]),
                    "summa":          float(dr["summa"]),
                    "uspeshka":       int(dr.get("uspeshka", 0)),
                    "uspeshka_summa": float(dr.get("uspeshka_summa", 0)),
                }

            # ── Build output rows ─────────────────────────────────────────────
            rows_out: List[Dict] = []
            for kpi in kpi_rows:
                code = str(kpi["staff_code"]).strip()
                staff_info = (
                    staff_by_code.get(code)
                    or staff_by_code.get(code.lstrip("0"))
                    or staff_by_code.get(code.zfill(4))
                )
                if not staff_info:
                    continue
                dept = staff_info["group"]
                if group and dept.upper() != group.upper():
                    continue
                consul         = int(kpi["consul"])
                zakas          = int(kpi["zakas"])
                dumka          = int(kpi["dumka"])
                summa          = float(kpi["summa"])
                uspeshka       = int(kpi.get("uspeshka", 0))
                uspeshka_summa = float(kpi.get("uspeshka_summa", 0))
                conv           = round(zakas / consul * 100, 1) if consul else 0.0
                rows_out.append({
                    "code":           code,
                    "name":           staff_info["full_name"],
                    "group":          dept,
                    "consul":         consul,
                    "zakas":          zakas,
                    "dumka":          dumka,
                    "summa":          int(summa),
                    "uspeshka":       uspeshka,
                    "uspeshka_summa": int(uspeshka_summa),
                    "conversion":     conv,
                    "daily":          daily_map.get(code, {}),
                })

            rows_out.sort(key=lambda x: (-x["summa"], x["name"]))
            for i, r in enumerate(rows_out, 1):
                r["num"] = i

            groups_out: Dict[str, List] = {}
            for r in rows_out:
                g = r["group"] or "—"
                groups_out.setdefault(g, []).append(r)

            all_consul         = sum(r["consul"]         for r in rows_out)
            all_zakas          = sum(r["zakas"]          for r in rows_out)
            all_dumka          = sum(r["dumka"]           for r in rows_out)
            all_summa          = sum(r["summa"]          for r in rows_out)
            all_uspeshka       = sum(r["uspeshka"]       for r in rows_out)
            all_uspeshka_summa = sum(r["uspeshka_summa"] for r in rows_out)
            avg_conv           = round(all_zakas / all_consul * 100, 1) if all_consul else 0.0

            return {
                "month":            month,
                "date_from":        date_from,
                "date_to":          date_to,
                "total_consul":     all_consul,
                "total_zakas":      all_zakas,
                "total_dumka":      all_dumka,
                "total_summa":      all_summa,
                "total_uspeshka":       all_uspeshka,
                "total_uspeshka_summa": all_uspeshka_summa,
                "avg_conversion":   avg_conv,
                "groups":           groups_out,
            }

        except Exception as exc:
            print(f"[DASHBOARD] monthly-report error: {traceback.format_exc()}")
            return {"error": str(exc), "month": month}

    # ── XLSX Export endpoint ──────────────────────────────────────────────────
    @router.get("/api/dashboard/export", tags=["dashboard"])
    def dashboard_export(
        request:  Request,
        date_from: str = Query(default=""),
        date_to:   str = Query(default=""),
        group:     str = Query(default=""),
    ):
        import traceback
        if not _user(request):
            from fastapi import HTTPException
            raise HTTPException(status_code=401, detail="Unauthorized")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                          Side)
            from openpyxl.utils import get_column_letter

            today = date.today().strftime("%Y-%m-%d")
            if not date_from: date_from = today
            if not date_to:   date_to   = today

            # ── Re-use cached stats (triggers a fetch if needed) ─────────────
            stats = dashboard_stats(date_from=date_from, date_to=date_to, group=group, force=0)

            # ── Style helpers ──────────────────────────────────────────────
            HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
            HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
            SUMM_FILL = PatternFill("solid", fgColor="0D1829")
            SUMM_FONT = Font(bold=True, color="93C5FD", size=10)
            GRP_FILL  = PatternFill("solid", fgColor="163352")
            GRP_FONT  = Font(bold=True, color="BAE6FD", size=10)
            THIN = Side(style="thin", color="334155")
            BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            CENTER = Alignment(horizontal="center", vertical="center")
            LEFT   = Alignment(horizontal="left",   vertical="center")

            def hdr(ws, row, cols):
                """Write a styled header row."""
                for c, val in enumerate(cols, 1):
                    cell = ws.cell(row=row, column=c, value=val)
                    cell.fill   = HDR_FILL
                    cell.font   = HDR_FONT
                    cell.border = BORDER
                    cell.alignment = CENTER

            def autofit(ws):
                """Set column widths based on max content length."""
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

            wb = Workbook()
            wb.remove(wb.active)  # remove default sheet

            # ═════════════════════════════════════════════════════════════════
            # Sheet 1: Сводка (Summary)
            # ═════════════════════════════════════════════════════════════════
            ws1 = wb.create_sheet("Сводка")
            ws1.sheet_view.showGridLines = False
            summary_rows = [
                ("Период",          f"{date_from}  –  {date_to}"),
                ("Группа фильтр",  group or "Все"),
                ("Консультации",     stats["total_consul"]),
                ("Заказы",           stats["total_zakas"]),
                ("Отказы",           stats["total_otkaz"]),
                ("Раздумья",        stats["total_dumka"]),
                ("Сумма заказов",  stats["total_summa"]),
                ("Конверсия",       f"{stats['avg_conversion']}%"),
            ]
            for r, (label, val) in enumerate(summary_rows, 1):
                lc = ws1.cell(row=r, column=1, value=label)
                lc.font = HDR_FONT; lc.fill = HDR_FILL; lc.border = BORDER; lc.alignment = LEFT
                vc = ws1.cell(row=r, column=2, value=val)
                vc.font = Font(bold=True, color="E2E8F0", size=11)
                vc.fill = PatternFill("solid", fgColor="151F32")
                vc.border = BORDER; vc.alignment = LEFT
            ws1.column_dimensions["A"].width = 22
            ws1.column_dimensions["B"].width = 24

            # ═════════════════════════════════════════════════════════════════
            # Sheet 2: Сотрудники (Staff KPI by group)
            # ═════════════════════════════════════════════════════════════════
            ws2 = wb.create_sheet("Сотрудники")
            ws2.sheet_view.showGridLines = False
            STAFF_COLS = ["#", "Отдел", "Код", "Сотрудник",
                          "Olag (Consul)", "Zakaz Soni", "Zakaz Konv.%",
                          "Qilingan Summa", "Uspeshka Summas", "Uspeshka Konv.%"]
            row_num = 1
            hdr(ws2, row_num, STAFF_COLS)
            row_num += 1
            for g_name, g_rows in stats.get("groups", {}).items():
                # Group header row
                gc = ws2.cell(row=row_num, column=1, value=f"Отдел {g_name}")
                gc.fill = GRP_FILL; gc.font = GRP_FONT; gc.alignment = LEFT; gc.border = BORDER
                ws2.merge_cells(start_row=row_num, start_column=1,
                                end_row=row_num,   end_column=len(STAFF_COLS))
                for c2 in range(2, len(STAFF_COLS) + 1):
                    ws2.cell(row=row_num, column=c2).fill = GRP_FILL
                    ws2.cell(row=row_num, column=c2).border = BORDER
                row_num += 1
                for r in g_rows:
                    vals = [r.get("num",""), g_name, r["code"], r["name"],
                            r["consul"], r["zakas"], r.get("zakaz_conv", 0),
                            r["summa"], r.get("uspeshka_summa", 0), r.get("uspeshka_conv", 0)]
                    for c2, v in enumerate(vals, 1):
                        cell = ws2.cell(row=row_num, column=c2, value=v)
                        cell.border = BORDER
                        cell.alignment = CENTER if c2 != 4 else LEFT
                        if c2 in (8, 9):  # summa columns
                            cell.number_format = "#,##0"
                        if c2 in (7, 10):  # conversion columns
                            cell.font = Font(bold=True, color=(
                                "4ADE80" if (v or 0) >= 50 else
                                "FACC15" if (v or 0) >= 25 else "F87171"))
                    row_num += 1
            autofit(ws2)

            # ── Serialize to bytes ───────────────────────────────────────────────
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            fname_label = f"{date_from}_{date_to}" + (f"_Группа{group}" if group else "")
            filename = f"KPI_{fname_label}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except Exception as exc:
            print(f"[DASHBOARD] Export error: {traceback.format_exc()}")
            return {"error": str(exc)}

    return router


# ─────────────────────────────────────────────────────────────────────────────
# Login page
# ─────────────────────────────────────────────────────────────────────────────

_LOGIN_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Вход — KPI Dashboard</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    body { background:#0b1120; font-family:'Inter',system-ui,sans-serif; }
    input { background:#ffffff !important; border:1px solid #2d3f5a; color:#0f172a !important;
            -webkit-text-fill-color:#0f172a !important;
            color-scheme: light !important;
            border-radius:8px; padding:10px 14px; width:100%; font-size:14px;
            outline:none; box-sizing:border-box; }
    input::placeholder { color:#94a3b8 !important; opacity:1; }
    input:focus, input:active, input:hover {
      background:#ffffff !important;
      color:#0f172a !important;
      -webkit-text-fill-color:#0f172a !important;
      color-scheme: light !important;
      border-color:#3b82f6 !important; }
    input:-webkit-autofill,
    input:-webkit-autofill:focus,
    input:-webkit-autofill:hover,
    input:-webkit-autofill:active {
      -webkit-box-shadow: 0 0 0 1000px #ffffff inset !important;
      -webkit-text-fill-color: #0f172a !important;
      color-scheme: light !important; }
    .btn-login { background:#2563eb; color:#fff; border-radius:8px; padding:11px;
                 width:100%; font-size:14px; font-weight:600; cursor:pointer;
                 border:none; transition:background .15s; }
    .btn-login:hover { background:#1d4ed8; }
    .err { background:#7f1d1d40; color:#fca5a5; border:1px solid #7f1d1d70;
           border-radius:8px; padding:10px 14px; font-size:13px; }
  </style>
</head>
<body class="min-h-screen flex items-center justify-center p-4">
  <div style="width:100%;max-width:380px">
    <div class="flex items-center gap-3 justify-center mb-8">
      <div style="width:40px;height:40px;background:#2563eb22;border:1px solid #2563eb44;
                  border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:20px">📊</div>
      <div>
        <div style="color:#f1f5f9;font-weight:700;font-size:18px">Staff KPI Dashboard</div>
        <div style="color:#475569;font-size:12px">amoCRM — реальное время</div>
      </div>
    </div>
    <div style="background:#151f32;border:1px solid #1e2d45;border-radius:14px;padding:28px 32px">
      <h2 style="color:#e2e8f0;font-weight:600;margin-bottom:22px;font-size:15px">Вход в систему</h2>
      {error_block}
      <form method="post" action="/login">
        <div style="margin-bottom:14px">
          <label style="color:#64748b;font-size:11px;text-transform:uppercase;
                        letter-spacing:.06em;display:block;margin-bottom:6px">Логин</label>
          <input type="text" name="username" autocomplete="username"
                 style="background:#ffffff !important;color:#0f172a !important;color-scheme:light"
                 placeholder="Введите логин" required />
        </div>
        <div style="margin-bottom:22px">
          <label style="color:#64748b;font-size:11px;text-transform:uppercase;
                        letter-spacing:.06em;display:block;margin-bottom:6px">Пароль</label>
          <input type="password" name="password" autocomplete="current-password"
                 style="background:#ffffff !important;color:#0f172a !important;color-scheme:light"
                 placeholder="Введите пароль" required />
        </div>
        <button type="submit" class="btn-login">Войти</button>
      </form>
    </div>
    <p style="color:#334155;font-size:11px;text-align:center;margin-top:16px">
      amoCRM → Google Sheets &nbsp;·&nbsp; KPI Dashboard
    </p>
  </div>
  <script>
    // Force visible text on all inputs — overrides any browser/autofill override
    function fixInputs() {
      document.querySelectorAll('input').forEach(function(el) {
        el.style.setProperty('color', '#0f172a', 'important');
        el.style.setProperty('background-color', '#ffffff', 'important');
        el.style.setProperty('opacity', '1', 'important');
      });
    }
    fixInputs();
    // Re-apply after autofill kicks in (Chrome fires it ~100ms after load)
    setTimeout(fixInputs, 200);
    setTimeout(fixInputs, 600);
    document.querySelectorAll('input').forEach(function(el) {
      el.addEventListener('animationstart', fixInputs);
      el.addEventListener('change', fixInputs);
    });
  </script>
</body>
</html>"""

# ─────────────────────────────────────────────────────────────────────────────
# Self-contained HTML dashboard page (Manager KPI — 10-column table)
# ─────────────────────────────────────────────────────────────────────────────

_DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Manager KPI Dashboard</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    body { background:#0b1120; color:#e2e8f0; font-family:'Inter',system-ui,sans-serif; }

    .card  { background:#151f32; border:1px solid #1e2d45; border-radius:14px; }

    /* ── Group header accent colours ── */
    .ghdr-a { background:linear-gradient(90deg,#1d4ed820,#0b1120 80%); border-left:3px solid #3b82f6; }
    .ghdr-b { background:linear-gradient(90deg,#15803d20,#0b1120 80%); border-left:3px solid #22c55e; }
    .ghdr-c { background:linear-gradient(90deg,#9d174d20,#0b1120 80%); border-left:3px solid #ec4899; }
    .ghdr-d { background:linear-gradient(90deg,#92400e20,#0b1120 80%); border-left:3px solid #f59e0b; }
    .ghdr-baza { background:linear-gradient(90deg,#5b21b620,#0b1120 80%); border-left:3px solid #a78bfa; }
    .ghdr-def  { background:linear-gradient(90deg,#33415520,#0b1120 80%); border-left:3px solid #64748b; }

    /* ── Table ── */
    .tbl { border-collapse:collapse; width:100%; }
    .tbl th { background:#0d1829; color:#64748b; font-size:10px; text-transform:uppercase;
              letter-spacing:.06em; padding:9px 10px; white-space:nowrap; position:sticky;
              top:0; z-index:1; cursor:pointer; user-select:none; }
    .tbl th:hover { color:#94a3b8; }
    .tbl th .si { opacity:.3; margin-left:2px; font-size:9px; }
    .tbl th.sa .si::after { content:'▲'; opacity:1; }
    .tbl th.sd .si::after { content:'▼'; opacity:1; }
    .tbl th.sa .si, .tbl th.sd .si { opacity:1; color:#3b82f6; }
    .tbl td { padding:7px 10px; font-size:12px; border-bottom:1px solid rgba(30,45,69,.7); }
    .tbl tr:last-child td { border-bottom:none; }
    .tbl tbody tr:hover td { background:rgba(59,130,246,.06); }
    .tbl tfoot td { background:#08101c; border-top:1px solid #1e2d45; font-size:12px; font-weight:700; }

    /* ── Conv bar ── */
    .cb-wrap { display:flex; align-items:center; gap:5px; justify-content:flex-end; }
    .cb-bg   { width:36px; height:4px; background:#1e2d45; border-radius:2px; flex-shrink:0; }
    .cb-fill { height:100%; border-radius:2px; }
    .ch { color:#4ade80; font-weight:700; }
    .cm { color:#facc15; font-weight:700; }
    .cl { color:#f87171; font-weight:700; }

    /* ── Badges ── */
    .badge-a    { background:#1d4ed830; color:#93c5fd; border:1px solid #1d4ed870; }
    .badge-b    { background:#15803d30; color:#86efac; border:1px solid #15803d70; }
    .badge-c    { background:#9d174d30; color:#f9a8d4; border:1px solid #9d174d70; }
    .badge-d    { background:#92400e30; color:#fcd34d; border:1px solid #92400e70; }
    .badge-baza { background:#5b21b630; color:#c4b5fd; border:1px solid #5b21b670; }
    .badge-def  { background:#33415530; color:#94a3b8; border:1px solid #33415570; }
    .badge      { display:inline-block; padding:2px 9px; border-radius:999px; font-size:11px; font-weight:700; }

    /* ── Status badges ── */
    .live-badge { background:#7f1d1d50; color:#fca5a5; border:1px solid #7f1d1d; border-radius:6px;
                  padding:3px 9px; font-size:11px; font-weight:700; }
    .snap-badge { background:#14532d50; color:#86efac; border:1px solid #14532d; border-radius:6px;
                  padding:3px 9px; font-size:11px; font-weight:700; }

    /* ── Buttons ── */
    .btn { padding:6px 14px; border-radius:8px; font-size:13px; font-weight:600; cursor:pointer;
           transition:all .15s; border:1px solid transparent; }
    .btn-primary { background:#2563eb; color:#fff; border-color:#2563eb; }
    .btn-primary:hover { background:#1d4ed8; }
    .btn-outline { background:transparent; border:1px solid #1e2d45; color:#64748b; }
    .btn-outline:hover { border-color:#3b82f6; color:#e2e8f0; }
    .btn-preset { padding:5px 11px; border-radius:6px; font-size:12px; font-weight:500; cursor:pointer;
                  background:transparent; border:1px solid #1e2d45; color:#64748b; transition:all .15s; }
    .btn-preset:hover { border-color:#3b82f6; color:#cbd5e1; }
    .btn-preset.active { background:#1e3a5f; border-color:#2563eb; color:#93c5fd; }
    .btn-active      { background:#2563eb !important; color:#fff !important; border-color:#2563eb !important; }
    .btn-active-a    { background:#1d4ed8 !important; color:#bfdbfe !important; border-color:#1d4ed8 !important; }
    .btn-active-b    { background:#15803d !important; color:#bbf7d0 !important; border-color:#15803d !important; }
    .btn-active-c    { background:#9d174d !important; color:#fce7f3 !important; border-color:#9d174d !important; }
    .btn-active-d    { background:#92400e !important; color:#fde68a !important; border-color:#92400e !important; }
    .btn-active-baza { background:#5b21b6 !important; color:#ede9fe !important; border-color:#5b21b6 !important; }

    /* ── Summary cards ── */
    .scard { border-radius:14px; padding:16px 18px; display:flex; align-items:center; gap:12px; }
    .scard-icon { width:40px; height:40px; border-radius:10px; display:flex; align-items:center;
                  justify-content:center; font-size:18px; flex-shrink:0; }
    .scard-val { font-size:22px; font-weight:700; line-height:1.1; font-variant-numeric:tabular-nums; }
    .scard-lbl { font-size:11px; color:#64748b; margin-top:2px; }

    /* ── Skeleton ── */
    .skeleton { background:linear-gradient(90deg,#151f32 25%,#1e2d45 50%,#151f32 75%);
                background-size:200% 100%; animation:shimmer 1.4s infinite; border-radius:8px; }
    @keyframes shimmer { 0%{background-position:200% 0} 100%{background-position:-200% 0} }
    .spinner { border:3px solid #1e2d45; border-top-color:#3b82f6; border-radius:50%;
               width:22px; height:22px; animation:spin .7s linear infinite; }
    @keyframes spin { to { transform:rotate(360deg); } }

    /* ── Rank medals ── */
    .rk1 { color:#facc15; } .rk2 { color:#94a3b8; } .rk3 { color:#c2722a; }

    /* ── Inputs ── */
    input[type="date"], input[type="text"] {
      background:#151f32; border:1px solid #1e2d45; color:#e2e8f0;
      border-radius:8px; padding:7px 10px; font-size:13px;
    }
    input[type="date"]:focus, input[type="text"]:focus { outline:none; border-color:#3b82f6; }

    /* ── Grid ── */
    #gc { display:grid; grid-template-columns:repeat(3,1fr); gap:14px; align-items:start; }
    #gc.g1 { grid-template-columns:1fr; }
    #gc.g2 { grid-template-columns:repeat(2,1fr); }
    #gc.g3 { grid-template-columns:repeat(3,1fr); }
    .gc-card.hidden-gc { display:none; }

    .tbl-scroll { max-height:480px; overflow-y:auto; }
    ::-webkit-scrollbar { width:4px; height:4px; }
    ::-webkit-scrollbar-track { background:#0b1120; }
    ::-webkit-scrollbar-thumb { background:#1e2d45; border-radius:2px; }
    ::-webkit-scrollbar-thumb:hover { background:#334155; }
  </style>
</head>
<body class="min-h-screen p-4 md:p-6">

  <!-- ── Header ── -->
  <div class="flex items-start justify-between mb-5 gap-4 flex-wrap">
    <div>
      <div class="flex items-center gap-2.5 mb-0.5">
        <div class="w-8 h-8 rounded-lg bg-blue-600/20 flex items-center justify-content:center text-blue-400 text-lg flex items-center justify-center">📊</div>
        <h1 class="text-xl font-bold text-white tracking-tight">Manager KPI Dashboard</h1>
      </div>
      <p class="text-slate-500 text-xs ml-10">amoCRM — данные из снепшотов + live kpi_events</p>
    </div>
    <div class="flex items-center gap-2 flex-wrap justify-end">
      <span id="data-badge"></span>
      <div id="spinner" class="spinner hidden"></div>
      <span id="last-upd" class="text-slate-600 text-xs hidden sm:inline"></span>
      <button id="btn-refresh" class="btn btn-outline text-xs py-1.5" onclick="loadStats(true)">↻ Обновить</button>
      <button class="btn btn-outline text-xs py-1.5" style="border-color:#22c55e55;color:#86efac" onclick="exportXlsx()">↓ XLSX</button>
      <a href="/logout" class="btn btn-outline text-xs py-1.5" style="border-color:#47556944;color:#64748b;text-decoration:none">→ Выйти</a>
      <label class="flex items-center gap-1.5 text-slate-500 text-xs cursor-pointer select-none">
        <input type="checkbox" id="auto-ref" class="accent-blue-500" onchange="toggleAutoRef()" />
        Авто 60 с
      </label>
    </div>
  </div>

  <!-- ── Filters ── -->
  <div class="card p-4 mb-5 flex flex-wrap items-end gap-3">
    <div>
      <label class="block text-[10px] text-slate-500 mb-1 uppercase tracking-wide">От</label>
      <input type="date" id="f-from" style="width:140px" />
    </div>
    <div>
      <label class="block text-[10px] text-slate-500 mb-1 uppercase tracking-wide">До</label>
      <input type="date" id="f-to" style="width:140px" />
    </div>
    <div class="flex gap-1.5 self-end pb-0.5">
      <button class="btn-preset active" onclick="presetDay(this)">Сегодня</button>
      <button class="btn-preset" onclick="presetWeek(this)">Неделя</button>
      <button class="btn-preset" onclick="presetMonth(this)">Месяц</button>
    </div>
    <button class="btn btn-primary self-end" onclick="loadStats()">Применить</button>

    <div class="w-px h-7 bg-slate-700/60 self-end hidden sm:block"></div>

    <div class="self-end">
      <label class="block text-[10px] text-slate-500 mb-1 uppercase tracking-wide">Группа</label>
      <div id="grp-btns" class="flex gap-1 flex-wrap">
        <button class="btn btn-outline btn-active text-xs py-1" data-g="" onclick="setGrp(this,'')">Все</button>
        <button class="btn btn-outline text-xs py-1" data-g="A"    onclick="setGrp(this,'A')">A</button>
        <button class="btn btn-outline text-xs py-1" data-g="B"    onclick="setGrp(this,'B')">B</button>
        <button class="btn btn-outline text-xs py-1" data-g="C"    onclick="setGrp(this,'C')">C</button>
        <button class="btn btn-outline text-xs py-1" data-g="D"    onclick="setGrp(this,'D')">D</button>
        <button class="btn btn-outline text-xs py-1" data-g="Baza" onclick="setGrp(this,'Baza')">Baza</button>
      </div>
    </div>

    <div class="flex-1 min-w-[150px] self-end">
      <label class="block text-[10px] text-slate-500 mb-1 uppercase tracking-wide">Поиск сотрудника</label>
      <input type="text" id="f-staff" placeholder="Имя или код…" oninput="filterStaff()" style="width:100%" />
    </div>
  </div>

  <!-- ── Summary Cards ── -->
  <div id="sum-area" class="grid grid-cols-2 sm:grid-cols-3 lg:grid-cols-6 gap-3 mb-5">
    <div class="skeleton h-20 rounded-xl"></div>
    <div class="skeleton h-20 rounded-xl"></div>
    <div class="skeleton h-20 rounded-xl"></div>
    <div class="skeleton h-20 rounded-xl"></div>
    <div class="skeleton h-20 rounded-xl"></div>
    <div class="skeleton h-20 rounded-xl"></div>
  </div>

  <!-- ── Group Tables ── -->
  <div id="gc" class="mb-5">
    <div id="sk-grid" class="grid grid-cols-2 md:grid-cols-4 gap-3">
      <div class="skeleton h-64 rounded-xl"></div>
      <div class="skeleton h-64 rounded-xl"></div>
      <div class="skeleton h-64 rounded-xl"></div>
      <div class="skeleton h-64 rounded-xl"></div>
    </div>
  </div>

  <!-- ── Error ── -->
  <div id="err-banner" class="hidden mb-4 card p-4 flex items-start gap-3"
       style="border-color:#7f1d1d;background:rgba(127,29,29,.2)">
    <span class="text-red-400 text-base flex-shrink-0">⚠</span>
    <div class="flex-1 min-w-0">
      <div class="text-red-300 font-semibold text-sm">Ошибка загрузки</div>
      <div id="err-text" class="text-red-400/80 text-xs mt-0.5 break-all"></div>
    </div>
    <button onclick="clearErr()" class="text-red-500 hover:text-red-300">✕</button>
  </div>

  <div id="empty-msg" class="hidden text-center text-slate-600 py-20 text-base">
    Нет данных за выбранный период.
  </div>

  <p class="text-center text-slate-700 text-xs mt-8 pb-4">
    amoCRM → Google Sheets &nbsp;·&nbsp; KPI Dashboard
  </p>

<script>
// ── State ─────────────────────────────────────────────────────────────────────
let activeGrp = '';
let autoTimer = null;
let sortState = {};   // tableId → {col, dir}

// ── Column definitions ─────────────────────────────────────────────────────────
// Columns: # | Код | Сотрудник | Olag | Zakaz Soni | Zakaz Konv.% | Qilingan Summa | Uspeshka Summas | Uspeshka Konv.%
const COLS = [
  { key:'num',           label:'#',              align:'left',  fmt: v => v },
  { key:'code',          label:'Код',            align:'right', fmt: v => `<span class="text-slate-500 text-[10px]">${v}</span>` },
  { key:'name',          label:'Сотрудник',      align:'left',  fmt: v => `<span class="text-slate-200 font-medium">${v}</span>` },
  { key:'consul',        label:'Olag',           align:'right', fmt: v => `<span class="text-blue-300 font-semibold">${v}</span>` },
  { key:'zakas',         label:'Zakaz Soni',     align:'right', fmt: v => `<span class="text-green-400 font-semibold">${v}</span>` },
  { key:'zakaz_conv',    label:'Zakaz Konv.%',   align:'right', fmt: v => convBar(v) },
  { key:'summa',         label:'Qilingan Summa', align:'right', fmt: v => `<span class="text-yellow-400 font-semibold">${fmtMoney(v)}</span>` },
  { key:'uspeshka_summa',label:'Uspeshka Summas',align:'right', fmt: v => `<span class="text-emerald-400 font-semibold">${fmtMoney(v)}</span>` },
  { key:'uspeshka_conv', label:'Uspeshka Konv.%',align:'right', fmt: v => convBar(v) },
];

// ── Init ─────────────────────────────────────────────────────────────────────
(function init() {
  const today = new Date().toISOString().slice(0,10);
  document.getElementById('f-from').value = today;
  document.getElementById('f-to').value   = today;
  loadStats();
})();

// ── Date presets ──────────────────────────────────────────────────────────────
function clearPresets() { document.querySelectorAll('.btn-preset').forEach(b => b.classList.remove('active')); }
function presetDay(b)  { clearPresets(); b.classList.add('active'); const t=isoToday(); setDates(t,t); loadStats(); }
function presetWeek(b) {
  clearPresets(); b.classList.add('active');
  const t=new Date(); const day=t.getDay()||7;
  const mn=new Date(t); mn.setDate(t.getDate()-day+1);
  setDates(mn.toISOString().slice(0,10), isoToday()); loadStats();
}
function presetMonth(b) {
  clearPresets(); b.classList.add('active');
  const t=new Date(); const mn=new Date(t.getFullYear(),t.getMonth(),1);
  setDates(mn.toISOString().slice(0,10), isoToday()); loadStats();
}
function isoToday() { return new Date().toISOString().slice(0,10); }
function setDates(f,t) {
  document.getElementById('f-from').value=f;
  document.getElementById('f-to').value=t;
}

// ── Group filter ────────────────────────────────────────────────────────────
const G_ACT = {'':'btn-active',A:'btn-active-a',B:'btn-active-b',C:'btn-active-c',D:'btn-active-d',Baza:'btn-active-baza'};
function setGrp(btn, g) {
  activeGrp = g;
  document.querySelectorAll('#grp-btns .btn').forEach(b =>
    b.classList.remove('btn-active','btn-active-a','btn-active-b','btn-active-c','btn-active-d','btn-active-baza'));
  btn.classList.add(G_ACT[g]||'btn-active');
  applyGrpVis();
}

function applyGrpVis() {
  const cards = document.querySelectorAll('.gc-card');
  let vis = 0;
  cards.forEach(c => {
    const show = !activeGrp || c.dataset.group.toUpperCase() === activeGrp.toUpperCase();
    c.classList.toggle('hidden-gc', !show);
    if (show) vis++;
  });
  const gc = document.getElementById('gc');
  gc.classList.remove('g1','g2','g3');
  if (vis===1) gc.classList.add('g1');
  else if (vis===2) gc.classList.add('g2');
  else if (vis===3) gc.classList.add('g3');
  filterStaff();
}

// ── Auto-refresh ──────────────────────────────────────────────────────────────
function toggleAutoRef() {
  clearInterval(autoTimer);
  if (document.getElementById('auto-ref').checked)
    autoTimer = setInterval(() => loadStats(false), 60_000);
}

// ── Load stats ────────────────────────────────────────────────────────────────
async function loadStats(force=false) {
  const from = document.getElementById('f-from').value;
  const to   = document.getElementById('f-to').value;
  document.getElementById('spinner').classList.remove('hidden');
  document.getElementById('btn-refresh').disabled = true;
  try {
    const fp = force ? '&force=1' : '';
    const res = await fetch(`/api/dashboard/stats?date_from=${from}&date_to=${to}${fp}`);
    if (!res.ok) { showErr(`HTTP ${res.status}`); return; }
    const data = await res.json();
    if (data.error) { showErr(data.error); return; }
    clearErr();
    renderBadge(data);
    renderSummary(data);
    renderGroups(data);
    applyGrpVis();
    filterStaff();
    const now = new Date().toLocaleTimeString('ru-RU',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
    const lu = document.getElementById('last-upd');
    lu.textContent = 'Обновлено '+now; lu.classList.remove('hidden');
    // Enable auto-refresh only for live data
    if (data.is_live && document.getElementById('auto-ref').checked) {
      clearInterval(autoTimer);
      autoTimer = setInterval(() => loadStats(false), 60_000);
    }
  } catch(e) { showErr(e.toString()); }
  finally {
    document.getElementById('spinner').classList.add('hidden');
    document.getElementById('btn-refresh').disabled = false;
  }
}

// ── Data-source badge ─────────────────────────────────────────────────────────
function renderBadge(data) {
  const el = document.getElementById('data-badge');
  el.innerHTML = data.is_live
    ? '<span class="live-badge">🔴 LIVE — kpi_events</span>'
    : '<span class="snap-badge">🟢 Snapshot</span>';
}

// ── Summary cards ─────────────────────────────────────────────────────────────
const SCARDS = [
  { lbl:'Olag (Consul)',      color:'#1e3a5f', ic:'💬', icBg:'#1d4ed820', icCol:'#60a5fa',  key:'total_consul' },
  { lbl:'Zakaz Soni',        color:'#14532d', ic:'✅', icBg:'#15803d20', icCol:'#4ade80',  key:'total_zakas' },
  { lbl:'Zakaz Konv.%',      color:'#2e1065', ic:'%',  icBg:'#5b21b620', icCol:'#c4b5fd',  key:'avg_zakaz_conv', sfx:'%' },
  { lbl:'Qilingan Summa',    color:'#422006', ic:'₸',  icBg:'#92400e20', icCol:'#fbbf24',  key:'total_summa' },
  { lbl:'Uspeshka Summas',   color:'#064e3b', ic:'🏆', icBg:'#06532520', icCol:'#34d399',  key:'total_uspeshka_summa' },
  { lbl:'Uspeshka Konv.%',   color:'#134e4a', ic:'📈', icBg:'#0f766e20', icCol:'#2dd4bf',  key:'avg_uspeshka_conv', sfx:'%' },
];

function renderSummary(data) {
  document.getElementById('sum-area').innerHTML = SCARDS.map(s => {
    const raw = data[s.key] ?? 0;
    const val = s.key === 'total_summa' || s.key === 'total_uspeshka_summa'
      ? fmtMoney(raw)
      : (raw + (s.sfx||''));
    return `<div class="scard" style="background:${s.color}30;border:1px solid ${s.color}80">
      <div class="scard-icon" style="background:${s.icBg};color:${s.icCol}">${s.ic}</div>
      <div><div class="scard-val" style="color:${s.icCol}">${val}</div>
           <div class="scard-lbl">${s.lbl}</div></div>
    </div>`;
  }).join('');
}

// ── Group tables ──────────────────────────────────────────────────────────────
const G_HDR   = {A:'ghdr-a',B:'ghdr-b',C:'ghdr-c',D:'ghdr-d',BAZA:'ghdr-baza'};
const G_BADGE = {A:'badge-a',B:'badge-b',C:'badge-c',D:'badge-d',BAZA:'badge-baza'};

function renderGroups(data) {
  const gc = document.getElementById('gc');
  const sk = document.getElementById('sk-grid');
  if (sk) sk.remove();
  gc.innerHTML = '';

  const groups = data.groups || {};
  const ORDER  = ['A','B','C','D','Baza'];
  const keys   = [...ORDER.filter(k=>groups[k]), ...Object.keys(groups).filter(k=>!ORDER.includes(k)).sort()];

  const emMsg = document.getElementById('empty-msg');
  if (!keys.length) { emMsg.classList.remove('hidden'); return; }
  emMsg.classList.add('hidden');

  for (const g of keys) {
    const rows    = groups[g];
    const gUp     = g.toUpperCase();
    const bCls    = G_BADGE[gUp] || 'badge-def';
    const hCls    = G_HDR[gUp]   || 'ghdr-def';
    const tid     = 'tbl-'+g;

    // Totals
    const tc  = rows.reduce((s,r)=>s+r.consul,0);
    const tz  = rows.reduce((s,r)=>s+r.zakas,0);
    const ts  = rows.reduce((s,r)=>s+r.summa,0);
    const tus = rows.reduce((s,r)=>s+(r.uspeshka_summa||0),0);
    const tzc = tc ? +(tz/tc*100).toFixed(1) : 0;
    const tuc = ts ? +(tus/ts*100).toFixed(1) : 0;

    const card = document.createElement('div');
    card.className   = 'card overflow-hidden gc-card';
    card.dataset.group = g;

    card.innerHTML = `
      <div class="${hCls} px-3 py-2.5 flex items-center justify-between">
        <span class="font-semibold text-white text-sm flex items-center gap-2">
          <span class="badge ${bCls}">${g}</span>
          Отдел ${g}
        </span>
        <span class="text-xs text-slate-400">${rows.length} чел.</span>
      </div>
      <div class="tbl-scroll">
        <table class="tbl" id="${tid}">
          <thead><tr>
            ${COLS.map((c,i)=>`<th class="text-${c.align}" data-col="${i}" data-key="${c.key}"
              >${c.label}<span class="si"></span></th>`).join('')}
          </tr></thead>
          <tbody>${buildRows(rows)}</tbody>
          <tfoot><tr>
            <td colspan="3" class="px-2.5 py-2 text-xs text-slate-400">Итого</td>
            <td class="px-2.5 py-2 text-right text-blue-400">${tc}</td>
            <td class="px-2.5 py-2 text-right text-green-400">${tz}</td>
            <td class="px-2.5 py-2 text-right">${convBar(tzc)}</td>
            <td class="px-2.5 py-2 text-right text-yellow-400">${fmtMoney(ts)}</td>
            <td class="px-2.5 py-2 text-right text-emerald-400">${fmtMoney(tus)}</td>
            <td class="px-2.5 py-2 text-right">${convBar(tuc)}</td>
          </tr></tfoot>
        </table>
      </div>`;

    card.querySelectorAll('th[data-col]').forEach(th => {
      th.addEventListener('click', () => sortTbl(tid, parseInt(th.dataset.col)));
    });
    gc.appendChild(card);
  }
}

function buildRows(rows) {
  return rows.map((r, idx) => {
    const rank = idx+1;
    const rk   = rank===1?'<span class="rk1">🥇</span>'
               : rank===2?'<span class="rk2">🥈</span>'
               : rank===3?'<span class="rk3">🥉</span>'
               : `<span class="text-slate-600 text-[10px]">${rank}</span>`;
    const cells = COLS.map((c,i) => {
      let raw = r[c.key] ?? 0;
      let html = (i===0) ? rk : c.fmt(raw);
      return `<td class="text-${c.align}">${html}</td>`;
    }).join('');
    return `<tr class="staff-row"
      data-name="${(r.name||'').toLowerCase()}"
      data-code="${r.code||''}"
      data-consul="${r.consul||0}"
      data-zakas="${r.zakas||0}"
      data-zakaz_conv="${r.zakaz_conv||0}"
      data-summa="${r.summa||0}"
      data-uspeshka_summa="${r.uspeshka_summa||0}"
      data-uspeshka_conv="${r.uspeshka_conv||0}">${cells}</tr>`;
  }).join('');
}

// ── Sort ──────────────────────────────────────────────────────────────────────
function sortTbl(tid, colIdx) {
  const table = document.getElementById(tid);
  if (!table) return;
  const prev = sortState[tid] || {col:-1, dir:'desc'};
  const dir  = (prev.col===colIdx && prev.dir==='desc') ? 'asc' : 'desc';
  sortState[tid] = {col:colIdx, dir};

  table.querySelectorAll('th').forEach(th => {
    th.classList.remove('sa','sd');
    if (parseInt(th.dataset.col)===colIdx) th.classList.add(dir==='asc'?'sa':'sd');
  });

  const tbody = table.querySelector('tbody');
  const trows = Array.from(tbody.querySelectorAll('tr.staff-row'));
  const key   = COLS[colIdx]?.key;

  trows.sort((a,b) => {
    if (colIdx===2) {
      const av=a.dataset.name, bv=b.dataset.name;
      return dir==='asc'?av.localeCompare(bv):bv.localeCompare(av);
    }
    const av=parseFloat(a.dataset[key]||0), bv=parseFloat(b.dataset[key]||0);
    return dir==='asc'?av-bv:bv-av;
  });
  trows.forEach(r=>tbody.appendChild(r));
}

// ── Staff search ──────────────────────────────────────────────────────────────
function filterStaff() {
  const q = document.getElementById('f-staff').value.toLowerCase().trim();
  document.querySelectorAll('.staff-row').forEach(tr => {
    const matchName = !q || (tr.dataset.name||'').includes(q) || (tr.dataset.code||'').includes(q);
    const card = tr.closest('.gc-card');
    const matchGrp  = !activeGrp || (card && card.dataset.group.toUpperCase()===activeGrp.toUpperCase());
    tr.style.display = (matchName && matchGrp) ? '' : 'none';
  });
}

// ── Export ────────────────────────────────────────────────────────────────────
function exportXlsx() {
  const from=document.getElementById('f-from').value, to=document.getElementById('f-to').value;
  const url = `/api/dashboard/export?date_from=${from}&date_to=${to}&group=${activeGrp}`;
  fetch(url)
    .then(r=>{ if(!r.ok) throw new Error('HTTP '+r.status); return r.blob(); })
    .then(blob=>{ const a=document.createElement('a'); a.href=URL.createObjectURL(blob);
                  a.download=`KPI_${from}_${to}.xlsx`; a.click(); })
    .catch(e=>showErr('Ошибка экспорта: '+e.message));
}

// ── Error ─────────────────────────────────────────────────────────────────────
function showErr(m) {
  document.getElementById('err-text').textContent=m;
  document.getElementById('err-banner').classList.remove('hidden');
}
function clearErr() { document.getElementById('err-banner').classList.add('hidden'); }

// ── Helpers ───────────────────────────────────────────────────────────────────
function fmtMoney(n) {
  if (!n) return '0';
  if (n>=1_000_000) return (n/1_000_000).toFixed(1).replace('.',',')+' млн';
  return (+n).toLocaleString('ru-RU');
}
function convCls(v) { return v>=50?'ch':v>=25?'cm':'cl'; }
function convBar(v) {
  const pct=Math.min(v,100), col=v>=50?'#4ade80':v>=25?'#facc15':'#f87171';
  return `<div class="cb-wrap"><span class="${convCls(v)}">${v}%</span>
    <div class="cb-bg"><div class="cb-fill" style="width:${pct}%;background:${col}"></div></div></div>`;
}
</script>
</body>
</html>"""
